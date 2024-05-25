'''
Function: This file extracts data from BEA datasets on the foreign affiliates of U.S.~multinational firms. 
Source: input/international folder, with files downloaded from https://www.bea.gov/international/di1usdop.
Steps: We extract files by year, and identify industry hierarchies using the indentation in the raw files. 
'''

import re
import glob
import os
import pandas as pd
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX

# Directory where the data are stored
dir_path    = '../../input/international/'
output_dir  = '../../output/other'
temp_dir    = '../../output/temp'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
if not os.path.exists(temp_dir):
    os.makedirs(temp_dir)
output_file_name = 'combined_international_tables.csv'

# Helper function to clean industry names extracted from PRN files
# This is a helpfer function for processing 1983 to 1997 PRN files
def clean_industry_name(name):
    match = re.search(r'[^a-zA-Z (),]', name)
    return name[:match.start()].strip() if match else name.strip()

# Helper function to convert a TXT file to PRN format
# This is a helper function for processing year 1993, where the data is in TXT format
def convert_txt_to_prn(input_file, output_file):
    with open(input_file, 'r') as file:
        content = file.readlines()
    formatted_content = ''.join(content)
    with open(output_file, 'w') as file:
        file.write(formatted_content)

# Function to process PRN files from 1983 to 1997
def process_prn_file(file_path, year):
    columns_common = [
        'industry_level', 'industry_name', 'nbraffiliates', 'totalassets',
        'sales', 'netincome', 'compensationemployees', 'nbremployees',
        'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
        'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate'
    ]

    extra_columns = [
        'industry_level', 'industry_name', 'placeholder_isi','nbraffiliates', 'totalassets','placeholder_net_property',
        'sales','netincome', 'compensationemployees', 'nbremployees','placeholder_usdirectinvestment',
        'placeholder_directinvestmentincome', 'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
        'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate'
    ]
    
    # Adjustments for different years
    columns_varied = {
        '1989': extra_columns,
        '1994': extra_columns,
        'default': columns_common
    }
    
    header_skip_map = {
        '1989': 6,
        'default': 5
    }
    
    header_skip = header_skip_map.get(year, header_skip_map['default'])
    columns = columns_varied.get(year, columns_varied['default'])
    all_columns = columns + ['year']
    
    extracted_data = []
    start_processing = False
    dash_count = 0
    last_industry_level_1_name = ""

    with open(file_path, 'r') as file:
        for line in file:
            if '----' in line:
                dash_count += 1
                start_processing = dash_count == header_skip
                if dash_count > header_skip:
                    break
                continue

            if start_processing:
                if line.strip() == '' or clean_industry_name(line).strip() == '':
                    continue
                industry_name = clean_industry_name(line)
                industry_level = len(re.match(r'^\s*', line).group(0)) // 2 + 1

                # Special handling for 1990
                if year == '1990':
                    industry_level -= 9 
                if industry_level == 1:
                    last_industry_level_1_name = industry_name
                elif industry_level == 2 or industry_level == 3:
                    industry_name = f'{last_industry_level_1_name}_{industry_name}'
                elif industry_level ==4 and industry_name.strip() == 'All industries':
                    industry_level = 0
                else:
                    continue
                numbers = re.findall(r'\(D\)|\d+(?:,\d{3})*(?:\.\d+)?|\s\s\.{4}\s|\s[A-Z]\s|\(\*\)', line)
                data_row = [industry_level, industry_name] +  [''] * (len(all_columns) - len(numbers) - 3) + numbers
                data_row.append(year)
                extracted_data.append(data_row)
    return pd.DataFrame(extracted_data, columns=all_columns)


# Helper function to process each Excel file from 1999 to 2008
def process_xlsx_file_1999_2008(file_path, year, columns, last_industry_name):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    processed_data = []

    for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
        row_values = [cell.value for cell in row]
        row_values = [str(item) for item in row_values]  # Convert all items to strings and strip whitespace

        if not row_values or 'Continued' in row_values[0]:
            continue
        row_values = row_values[:(len(columns) - 2)]  
        industry_name = row_values[0]
        if industry_name.isdigit():
            continue
        if row[0].value and row[0].value.strip() != '':
            indentation = row[0].alignment.indent
            industry_level = indentation + 1
            if industry_level == 1:
                last_industry_name['name'] = industry_name
            elif industry_level == 2 and last_industry_name['name']:
                industry_name = f"{last_industry_name['name']}_{industry_name}"
            elif industry_level == 3 and last_industry_name['name']:
                industry_name = f"{last_industry_name['name']}_{industry_name}"
            elif industry_level == 4 and industry_name.strip() == 'All industries':
                industry_level = 0
            processed_data.append([industry_level, industry_name] + row_values[1:] + [year])
    return pd.DataFrame(processed_data, columns=columns)

# Function to handle file conversions and processing from 1999 to 2008
def convert_and_process_files(folder, year, columns):
    last_industry_name = {'name': None}
    for xls_file in os.listdir(folder):
        if xls_file.endswith('.xls'):
            xls_path = os.path.join(folder, xls_file)
            xlsx_path = os.path.join(temp_dir, xls_file[:-3] + 'xlsx')  # Save in temp_dir
            if not os.path.exists(xlsx_path):
                converter = XLS2XLSX(xls_path)
                converter.to_xlsx(xlsx_path)

    xlsx_files = sorted([f for f in os.listdir(temp_dir) if f.endswith('.xlsx')], key=lambda x: x[-6:])
    cumulative_df = pd.DataFrame()
    for xlsx_file in xlsx_files:
        file_path = os.path.join(temp_dir, xlsx_file)
        data = process_xlsx_file_1999_2008(file_path, year, columns, last_industry_name)
        cumulative_df = pd.concat([cumulative_df, data], ignore_index=True)
        os.remove(file_path)  #Remove the xlsx file after processing
    return cumulative_df

# Helper function to process each Excel file from 2009 onwards
# This function is used for checking if industry_name contains digits
def contains_digits(cell_value):
    return any(char.isdigit() for char in cell_value)

# Function to process Excel files from 2009 onwards
def process_xlsx_file_2009_2018(file_path, year, sheet_name, columns):
    global last_industry_name
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    processed_data = []
    for row in sheet.iter_rows(min_row=2):
        row_values = [cell.value if cell.value is not None else '' for cell in row[:len(columns)]]
        row_values = [str(item).strip() for item in row_values]
        if 'Continued' not in row_values[0]:
            cell = row[0]
            if cell.value and cell.value.strip() != '' and not contains_digits(cell.value) and not cell.value.startswith('See'):
                indentation = cell.alignment.indent
                industry_name = cell.value.strip()
                industry_level = indentation + 1
                new_row = [industry_level] + row_values
                new_row = new_row[:len(columns) - 1]

                # Here we hardcode indentation to correct typo in original data
                if year == '2009' and industry_name.strip() in ['Agriculture, forestry, fishing, and hunting', 'Crop production','Animal production','Forestry and logging','Fishing, hunting, and trapping','Support activities for agriculture and forestry']:
                    industry_level -=1
                    new_row[0] = industry_level
                if industry_level == 1:
                    last_industry_name = industry_name
                elif (industry_level == 2 or industry_level == 3) and last_industry_name:
                    new_row[1] = f'{last_industry_name}_{industry_name}'
                elif industry_level == 4 and industry_name.strip() == 'All industries':
                    industry_level = new_row[0] = 0
                if industry_level <= 3:
                    processed_data.append(new_row + [year])
    return processed_data

def update_industry_name_1998(data):
    data['indentations'] = data['industry_name'].apply(lambda x: len(x) - len(x.lstrip(' ')))
    data['indentations'] = data['indentations'] // 3
    # Add indentations to the industry_level
    data['adjusted_industry_level'] = data['industry_level'] + data['indentations']

    # Remove dots from the industry_name
    filtered_data = data[(data['adjusted_industry_level'] <= 4)].copy()
    filtered_data.loc[:, 'industry_name'] = filtered_data['industry_name'].str.replace('.', '', regex=False).str.strip()

    # Select the required columns for the final output
    final_data = filtered_data[['adjusted_industry_level',
            'industry_name', 'nbraffiliates', 'totalassets',
            'sales', 'netincome', 'compensationemployees', 'nbremployees',
            'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
            'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate',
            'year'
        ]]

    # Initialize a variable to hold the last industry_name of industry_level == 1
    last_level_1_industry_name = ""

    # Iterate through rows to update industry_name based on the described rule
    for index, row in final_data.iterrows():
        if row['adjusted_industry_level'] == 1:
            last_level_1_industry_name = row['industry_name']
        elif row['adjusted_industry_level'] == 2 or row['adjusted_industry_level'] == 3:
            final_data.at[index, 'industry_name'] = f"{last_level_1_industry_name}_{row['industry_name']}"
        elif row['adjusted_industry_level'] == 4 and row['industry_name'].strip() == 'All industries':
            final_data.at[index, 'adjusted_industry_level'] = 0

    return_data = final_data.copy()
    return_data.rename(columns={'adjusted_industry_level': 'industry_level'}, inplace=True)
    return_data = return_data[return_data['industry_level'].isin([0, 1, 2, 3])]
    return  return_data.applymap(lambda x: x.strip() if isinstance(x, str) else x)

# Define the common column configurations for different years from 1998 to 2018
common_columns = {
    'standard': ['industry_level', 'industry_name', 'nbraffiliates', 'totalassets',
                 'sales', 'netincome', 'compensationemployees', 'nbremployees',
                 'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
                 'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate', 'year'],

    'with_placeholders': ['industry_level', 'industry_name', 'placeholder_isi', 'nbraffiliates', 'totalassets',
                          'placeholder_net_property', 'sales', 'netincome', 'compensationemployees', 'nbremployees',
                          'placeholder_usdirectinvestment', 'placeholder_directinvestmentincome',
                          'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
                          'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate', 'year'],

    'isi_placeholders': ['industry_level', 'industry_name', 'placeholder_isi', 'nbraffiliates', 'totalassets',
                         'sales', 'netincome', 'compensationemployees', 'nbremployees',
                         'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
                         'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate', 'year'],
    
    'most_placeholders': ['industry_level', 'industry_name', 'placeholder_isi', 'nbraffiliates', 'totalassets',
                          'placeholder_net_property', 'sales', 'netincome', 'placeholder_rnd', 'compensationemployees', 'nbremployees',
                            'placeholder_usdirectinvestment','placeholder_directinvestmentincome',
                          'byparentindustry_nbrUSparents', 'byparentindustry_totalassets_parent',
                          'byparentindustry_nbraffiliates', 'byparentindustry_totalassets_affiliate', 'year']
}
# Map years to their respective column configurations
year_settings = {
    '1998': common_columns['standard'],
    '1999': common_columns['with_placeholders'],
    '2000': common_columns['standard'],
    '2001': common_columns['standard'],
    '2002': common_columns['standard'],
    '2003': common_columns['standard'],
    '2004': common_columns['with_placeholders'],
    '2005': common_columns['isi_placeholders'],
    '2006': common_columns['isi_placeholders'],
    '2007': common_columns['isi_placeholders'],
    '2008': common_columns['isi_placeholders'],
    '2009': common_columns['with_placeholders'],
    '2010': common_columns['isi_placeholders'],
    '2011': common_columns['isi_placeholders'],
    '2012': common_columns['isi_placeholders'],
    '2013': common_columns['isi_placeholders']
}

# Process the PRN data from 1983 to 1997
input_file = os.path.join(dir_path, '1993', 'TAB02.TXT')
output_file = os.path.join(temp_dir, 'TAB02.prn')
# Function call to convert text to prn using the updated paths
convert_txt_to_prn(input_file, output_file)
results_1993 = process_prn_file(output_file, '1993')

# Filter out columns containing 'placeholder' for 1993 and save
filtered_results_1993 = results_1993.loc[:, ~results_1993.columns.str.contains('placeholder')]
filtered_results_1993.to_csv(os.path.join(output_dir, 'cumulative_data_1993.csv'), index=False)
os.remove(output_file) 

for folder in glob.glob(os.path.join(dir_path, '[0-9][0-9][0-9][0-9]')):
    year = os.path.basename(folder)
    if int(year) not in range (1983,1998):
        continue
    if year == '1993':
        continue
    prn_files = sorted(glob.glob(os.path.join(folder, '*.PRN')) + glob.glob(os.path.join(folder, '*.prn')), key=lambda x: x[-5])
    df_list = [process_prn_file(file_path, year) for file_path in prn_files]
    results = pd.concat(df_list, ignore_index=True)
    filtered_results = results.loc[:, ~results.columns.str.contains('placeholder')]
    filtered_results.to_csv(os.path.join(output_dir, f'cumulative_data_{year}.csv'), index=False)

for folder in glob.glob(os.path.join(dir_path, '[0-9][0-9][0-9][0-9]')):
    year = os.path.basename(folder)
    if int(year) not in range(1998, 2009):
        continue
    columns = year_settings[year]
    result_df = convert_and_process_files(folder, year, columns)
    result_df = result_df[result_df['industry_level'].isin([0, 1, 2, 3])]
    result_df = result_df[~result_df['industry_name'].str.contains(r'\d')]
    filtered_results = result_df.loc[:, ~result_df.columns.str.contains('placeholder')]
    if year == '1998':
        filtered_results = update_industry_name_1998(filtered_results)
    filtered_results.to_csv(os.path.join(output_dir, f'cumulative_data_{year}.csv'), index=False)

# Processing 2009 onward
for folder in glob.glob(os.path.join(dir_path, '[0-9][0-9][0-9][0-9]')):
    cumulative_df = pd.DataFrame()
    year = os.path.basename(folder)
    if int(year) not in range(2009,2014):
        continue
    columns = year_settings[year]
    for xls_file in os.listdir(folder):
        if xls_file.endswith('.xls'):
            xls_file_path = os.path.join(folder, xls_file)
            xlsx_file_path = os.path.join(temp_dir, 'converted.xlsx')
            x2x = XLS2XLSX(xls_file_path)
            x2x.to_xlsx(xlsx_file_path)
            workbook = load_workbook(filename=xlsx_file_path)
            tab_names = [tab.title for tab in workbook.worksheets if 'I.A 2' in tab.title or 'Table I.A2' in tab.title]
            sorted_tabs = sorted(tab_names, key=lambda x: int(re.search(r'\((\d+)\)', x).group(1)))
            for tab_name in sorted_tabs:
                file_data = process_xlsx_file_2009_2018(xlsx_file_path, year, tab_name, columns)
                cumulative_df = pd.concat([cumulative_df, pd.DataFrame(file_data, columns=columns)], ignore_index=True)
    
    cumulative_df = cumulative_df.loc[:, ~cumulative_df.columns.str.contains('placeholder')]
    cumulative_df.to_csv(f'{output_dir}/cumulative_data_{year}.csv', index=False)
    os.remove(xlsx_file_path) 

# Combine all the data into a single CSV file and remove the working files
# Get all CSV files in the directory
individual_year_files = [f for f in os.listdir(output_dir) if f.endswith('.csv')]
individual_year_files.sort(key=lambda x: x[-8:-4])
appended_data = pd.DataFrame()
for file in individual_year_files:
    if file.startswith('cumulative'):
        file_path = os.path.join(output_dir, file)
        data = pd.read_csv(file_path)
        appended_data = pd.concat([appended_data, data], ignore_index=True)
        # Clean the files after appending
        os.remove(file_path)

appended_data.to_csv(os.path.join(output_dir, output_file_name), index=False)
print(f"Data combined and saved to {output_file_name}.")
